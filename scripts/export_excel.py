"""Dump data/university.db to data/university.xlsx, one worksheet per table.

Exists so the agent's answers can be checked by hand: open the workbook, sort or
filter the sheet the query touched, and compare. The Overview sheet lists every
table with its row count and column types, mirroring describe_table().
"""

from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "university.db"
XLSX_PATH = ROOT / "data" / "university.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MAX_COL_WIDTH = 44


def _tables(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        "SELECT name FROM sqlite_master WHERE type = 'table'"
        " AND name NOT LIKE 'sqlite_%' ORDER BY name"
    ).fetchall()
    return [r[0] for r in rows]


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def _autosize(ws) -> None:
    # openpyxl has no measure pass, so width comes from the longest rendered value.
    for col in range(1, ws.max_column + 1):
        widest = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(widest + 2, MAX_COL_WIDTH)


def _write_table(wb: Workbook, conn: sqlite3.Connection, table: str) -> tuple[int, list[str]]:
    cur = conn.execute(f"SELECT * FROM {table}")  # table names come from sqlite_master
    columns = [d[0] for d in cur.description]

    ws = wb.create_sheet(title=table[:31])
    ws.append(columns)
    n = 0
    for row in cur:
        ws.append(list(row))
        n += 1

    _style_header(ws, len(columns))
    _autosize(ws)
    return n, columns


def _write_overview(wb: Workbook, conn: sqlite3.Connection, tables: list[str],
                    counts: dict[str, int]) -> None:
    ws = wb.create_sheet(title="Overview", index=0)
    ws.append(["Table", "Rows", "Column", "Type", "Not null", "Primary key", "References"])

    fks: dict[str, dict[str, str]] = {}
    for table in tables:
        fks[table] = {
            r[3]: f"{r[2]}({r[4]})"
            for r in conn.execute(f"PRAGMA foreign_key_list({table})")
        }

    for table in tables:
        first = True
        for col in conn.execute(f"PRAGMA table_info({table})"):
            _cid, name, ctype, notnull, _default, pk = col
            ws.append([
                table if first else "",
                counts[table] if first else "",
                name, ctype, "YES" if notnull else "", "YES" if pk else "",
                fks[table].get(name, ""),
            ])
            first = False

    _style_header(ws, 7)
    _autosize(ws)


def export() -> None:
    if not DB_PATH.exists():
        raise SystemExit(f"{DB_PATH} not found - run scripts/init_db.py first")

    conn = sqlite3.connect(DB_PATH)
    try:
        tables = _tables(conn)
        wb = Workbook()
        wb.remove(wb.active)  # drop the default empty sheet

        counts: dict[str, int] = {}
        for table in tables:
            counts[table], _cols = _write_table(wb, conn, table)

        _write_overview(wb, conn, tables, counts)
        wb.save(XLSX_PATH)
    finally:
        conn.close()

    print(f"Wrote {XLSX_PATH}")
    print(f"  {'Overview':<18} {len(tables):>6} tables")
    for table in tables:
        print(f"  {table:<18} {counts[table]:>6} rows")


if __name__ == "__main__":
    try:
        export()
    except Exception as exc:
        print(f"export_excel failed: {exc}", file=sys.stderr)
        raise SystemExit(1)
